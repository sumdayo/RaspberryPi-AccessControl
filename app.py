import time
import requests
import json
import os
import sys
import math
import threading
import calendar
import openpyxl
import zipfile # `BadZipFile` エラー対策

from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta, date
from collections import defaultdict
from smartcard.System import readers
from smartcard.util import toHexString
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv # 環境変数の読み込み用に追加した

load_dotenv() # .envファイルから環境変数を読み込む

# -- Flaskアプリケーションの設定 --
app = Flask(__name__)
# FlaskのSECRET_KEYを環境変数から読み込み
app.config['SECRET_KEY'] = os.getenv("FLASK_SECRET_KEY", "default_insecure_fallback") 
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///access_log.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# --- 削除と修正操作用のパスワード設定 ---
EDIT_PASSWORD = os.getenv("APP_EDIT_PASSWORD")
DELETE_PASSWORD = os.getenv("APP_DELETE_PASSWORD")

if not os.getenv("FLASK_SECRET_KEY") or not EDIT_PASSWORD:
    # 秘密鍵または認証パスワードがない場合は起動させない、より安全なロジック
    print("Error: 環境変数が設定されていません。アプリを停止します。")
    import sys
    sys.exit(1)

db = SQLAlchemy(app)

# --- データベースモデルの定義 ---
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    idm = db.Column(db.String(160), unique=True, nullable=False)
    name = db.Column(db.String(80), nullable=False)

    def __repr__(self):
        return f'<User {self.name} ({self.idm})>'

class AccessLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)
    status = db.Column(db.String(20), nullable=False)
    
    user = db.relationship('User', backref=db.backref('access_logs', lazy=True))

    def __repr__(self):
        return f'<AccessLog {self.user.name} {self.status} at {self.timestamp}>'

def read_felica_card_idm():
    try:
        available_readers = readers()
        if not available_readers:
            print("NFC Error: No smart card readers found. Is PaSoRi connected and pcscd running?")
            return None

        pasori_reader = next((r for r in available_readers if 'PaSoRi' in str(r) or 'FeliCa' in str(r)), None)

        if not pasori_reader:
            print("NFC Error: PaSoRi reader not found. Please check its name or connection.")
            return None

        connection = pasori_reader.createConnection()
        connection.connect()
        
        idm_apdu = [0xFF, 0xCA, 0x00, 0x00, 0x00]
        response, sw1, sw2 = connection.transmit(idm_apdu)
        
        if sw1 == 0x90 and sw2 == 0x00:
            idm = toHexString(response).replace(" ", "")
            return idm
        else:
            print(f"NFC Error: Failed to get IDm. SW: {hex(sw1)} {hex(sw2)}")
            return None

    except Exception as e:
        return None

# --- Discord ウェブフックURLを設定 ---
DISCORD_WEBHOOK_URL = "https://discordapp.com/api/webhooks/1393286247258128404/XjqQlaaFHl3Xfa3zLSuMpk97UR_zlX1uYRzBu3XBiyQPbpOH-exNAY98IN44CCd9oFew"

def send_discord_notification(username, event_type, success=True, details=None):
    if not DISCORD_WEBHOOK_URL:
        print("Discord ウェブフックURLが設定されていません。通知はスキップされます。")
        return

    current_time = datetime.now().strftime("%Y年%m月%d日 %H時%M分%S秒")
    title = ""
    description = ""
    color = 0

    if event_type == '入室':
        title = "アクセスイベント: 入室"
        description = f"✅ {current_time}: **{username}** が **入室** しました。"
        color = 65280
    elif event_type == '退室':
        title = "アクセスイベント: 退室"
        description = f"🚪 {current_time}: **{username}** が **退室** しました。"
        color = 3447003
    elif event_type == 'アクセス試行':
        title = "アクセスイベント: アクセス失敗"
        description = f"❌ {current_time}: **{username}** が **アクセス** に失敗しました。"
        color = 16711680
    elif event_type == 'ユーザー追加':
        title = "ユーザー管理: 新規ユーザー追加"
        description = f"➕ {current_time}: 新しいユーザー **{username}** が追加されました。\nIDm: `{details.get('idm', 'N/A')}`"
        color = 65280
    elif event_type == 'ユーザー更新':
        title = "ユーザー管理: ユーザー情報更新"
        description = f"✏️ {current_time}: ユーザー **{username}** の情報が更新されました。"
        if details:
            if 'old_name' in details and 'new_name' in details and details['old_name'] != details['new_name']:
                description += f"\n名前: `{details['old_name']}` -> `{details['new_name']}`"
            if 'old_idm' in details and 'new_idm' in details and details['old_idm'] != details['new_idm']:
                description += f"\nIDm: `{details['old_idm']}` -> `{details['new_idm']}`"
        color = 16776960
    elif event_type == 'ユーザー削除':
        title = "ユーザー管理: ユーザー削除"
        description = f"🗑️ {current_time}: ユーザー **{username}** が削除されました。"
        color = 16711680
    elif event_type == 'ログ削除':
        title = "ユーザー管理: ログ削除"
        description = f"🧹 {current_time}: ユーザー **{username}** の入退室ログがすべて削除されました。"
        color = 7829367
    else:
        title = "不明なイベント"
        description = f"{current_time}: 不明なイベントが発生しました。"
        color = 7829367

    payload = {
        "embeds": [
            {
                "title": title,
                "description": description,
                "color": color,
                "fields": [
                    {"name": "ユーザー名", "value": username, "inline": True},
                    {"name": "時刻", "value": current_time, "inline": True},
                    {"name": "結果", "value": "成功" if success else "失敗", "inline": True}
                ],
                "footer": {
                    "text": "Raspberry Pi アクセス制御システム"
                },
                "timestamp": datetime.utcnow().isoformat() + "Z"
            }
        ]
    }

    headers = {"Content-Type": "application/json"}
    try:
        response = requests.post(DISCORD_WEBHOOK_URL, data=json.dumps(payload), headers=headers)
        response.raise_for_status()
        print(f"Discord 通知を送信しました: {title} - {username}")
    except requests.exceptions.RequestException as e:
        print(f"Discord 通知の送信中にエラーが発生しました: {e}")
        print(f"レスポンス内容: {response.text if 'response' in locals() else 'N/A'}")


# --- カード読み取りと処理のメインループ（別スレッドで実行） ---
def card_reading_loop():
    while True:
        idm = read_felica_card_idm()
        if idm:
            with app.app_context():
                user = User.query.filter_by(idm=idm).first()
                if user:
                    last_log = AccessLog.query.filter_by(user_id=user.id)\
                                 .order_by(AccessLog.timestamp.desc()).first()
                    
                    if last_log and last_log.status == '入室':
                        new_status = '退室'
                    else:
                        new_status = '入室'
                    
                    log_entry = AccessLog(user_id=user.id, status=new_status)
                    db.session.add(log_entry)
                    db.session.commit()

                    print(f"Access recorded: {user.name} - {new_status}")
                    send_discord_notification(user.name, new_status, success=True)
                else:
                    print(f"Unknown card detected! IDm: {idm}")
                    send_discord_notification("不明なユーザー", "アクセス試行", success=False)
            
            time.sleep(5)
        time.sleep(0.5)

# --- 滞在時間計算ヘルパー関数 ---
def _calculate_stay_time_for_logs(logs):
    """
    与えられたログリストからユーザーごとの滞在時間を計算します。
    入室と退室がペアになっているセッションのみをカウントします。
    """
    user_sessions_data = defaultdict(lambda: {'name': 'Unknown', 'total_seconds': 0})
    current_entry_times = {} # {user_id: entry_datetime}

    for log in logs:
        user_id = log.user_id
        timestamp = log.timestamp
        status = log.status

        if status == '入室':
            current_entry_times[user_id] = timestamp
        elif status == '退室':
            if user_id in current_entry_times:
                entry_time = current_entry_times.pop(user_id)
                duration = timestamp - entry_time
                user_sessions_data[user_id]['total_seconds'] += duration.total_seconds()
            # else: 対応する入室がない退室ログは無視

    # ユーザー名を取得してデータに追加
    with app.app_context(): # データベース操作のためにコンテキストが必要
        for user_id in user_sessions_data.keys():
            user_obj = User.query.get(user_id)
            if user_obj:
                user_sessions_data[user_id]['name'] = user_obj.name
            else:
                user_sessions_data[user_id]['name'] = f"不明なユーザー (ID:{user_id})"

    # ランキング形式に整形し、秒数を時間:分:秒形式に変換
    ranking_data = []
    for user_id, data in user_sessions_data.items():
        total_seconds = data['total_seconds']
        hours = int(total_seconds // 3600)
        minutes = int((total_seconds % 3600) // 60)
        seconds = int(total_seconds % 60)
        formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"
        ranking_data.append({
            'user_id': user_id,
            'name': data['name'],
            'total_seconds': total_seconds,
            'formatted_time': formatted_time
        })

    # 合計滞在時間で降順にソート
    ranking_data.sort(key=lambda x: x['total_seconds'], reverse=True)
    
    return ranking_data

# --- 月間滞在時間計算関数 ---
def calculate_monthly_stay_time(year, month):
    start_date = datetime(year, month, 1)
    if month == 12:
        end_date = datetime(year + 1, 1, 1)
    else:
        end_date = datetime(year, month + 1, 1)

    logs = AccessLog.query.filter(
        AccessLog.timestamp >= start_date,
        AccessLog.timestamp < end_date
    ).order_by(AccessLog.user_id, AccessLog.timestamp).all()
    
    return _calculate_stay_time_for_logs(logs)

# --- 今週の滞在時間計算関数 ---
def calculate_weekly_stay_time():
    today = datetime.now()
    # 週の始まり（月曜日）を計算
    start_of_week = today - timedelta(days=today.weekday())
    start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
    end_of_week = start_of_week + timedelta(days=7)

    logs = AccessLog.query.filter(
        AccessLog.timestamp >= start_of_week,
        AccessLog.timestamp < end_of_week
    ).order_by(AccessLog.user_id, AccessLog.timestamp).all()

    return _calculate_stay_time_for_logs(logs)

# --- 今までの合計滞在時間計算関数 ---
def calculate_total_stay_time():
    logs = AccessLog.query.order_by(AccessLog.user_id, AccessLog.timestamp).all()
    return _calculate_stay_time_for_logs(logs)

# --- カレンダー表示用のアクセスサマリー取得関数 ---
def get_monthly_access_summary(year, month):
    start_date = datetime(year, month, 1).date()
    if month == 12:
        end_date = datetime(year + 1, 1, 1).date()
    else:
        end_date = datetime(year, month + 1, 1).date()

    # 該当月の全てのログを取得
    logs = AccessLog.query.filter(
        AccessLog.timestamp >= start_date,
        AccessLog.timestamp < end_date
    ).order_by(AccessLog.timestamp).all()

    access_summary = defaultdict(set) # 日付ごとにアクセスしたユーザー名を格納
    user_names_cache = {} # ユーザー名をキャッシュ

    with app.app_context():
        for log in logs:
            log_date = log.timestamp.date()
            if log_date >= start_date and log_date < end_date:
                if log.user_id not in user_names_cache:
                    user_obj = User.query.get(log.user_id)
                    if user_obj:
                        user_names_cache[log.user_id] = user_obj.name
                    else:
                        user_names_cache[log.user_id] = f"不明なユーザー (ID:{log.user_id})"
                access_summary[log_date].add(user_names_cache[log.user_id])
    
    # setをlistに変換してソート
    for date_key in access_summary:
        access_summary[date_key] = sorted(list(access_summary[date_key]))

    return access_summary


# --- Webアプリケーションのルート定義 ---

@app.route('/')
def index():
    access_logs = AccessLog.query.order_by(AccessLog.timestamp.desc()).limit(20).all() # 最新20件

    # カレンダー表示用の年と月を取得
    current_year = datetime.now().year
    current_month = datetime.now().month
    
    # クエリパラメータから年と月を取得（カレンダーナビゲーション用）
    calendar_year = request.args.get('calendar_year', type=int, default=current_year)
    calendar_month = request.args.get('calendar_month', type=int, default=current_month)

    # デバッグ用: テンプレートに渡す値を確認
    print(f"DEBUG: Rendering index.html with calendar_year={calendar_year}, calendar_month={calendar_month}")

    # カレンダーデータ
    cal = calendar.Calendar(firstweekday=calendar.SUNDAY) # 日曜日始まり
    month_calendar = cal.monthdatescalendar(calendar_year, calendar_month)
    access_summary = get_monthly_access_summary(calendar_year, calendar_month)

    # 今月のランキング
    monthly_ranking = calculate_monthly_stay_time(current_year, current_month)

    # 今週のランキング
    weekly_ranking = calculate_weekly_stay_time()

    # 今までの合計ランキング
    total_ranking = calculate_total_stay_time()

    return render_template('index.html', 
                           logs=access_logs,
                           monthly_ranking=monthly_ranking,
                           weekly_ranking=weekly_ranking,
                           total_ranking=total_ranking,
                           current_year=current_year, # 今月のランキング表示用
                           current_month=current_month, # 今月のランキング表示用
                           calendar_year=calendar_year, # カレンダー表示用
                           calendar_month=calendar_month, # カレンダー表示用
                           month_calendar=month_calendar, # カレンダーデータ
                           access_summary=access_summary, # 日ごとのアクセスサマリー
                           datetime=datetime # テンプレートでdatetimeオブジェクトを使用するため
                           )

@app.route('/users', methods=['GET', 'POST'])
def manage_users():
    if request.method == 'POST':
        action = request.form.get('action')
        user_id = request.form.get('user_id')
        password = request.form.get('password') # パスワードを取得

        if action == 'add':
            idm = request.form['idm'].strip()
            name = request.form['name'].strip()
            if idm and name:
                with app.app_context():
                    existing_user = User.query.filter_by(idm=idm).first()
                    if not existing_user:
                        new_user = User(idm=idm, name=name)
                        db.session.add(new_user)
                        db.session.commit()
                        flash(f'ユーザー "{name}" を追加しました。', 'success')
                        send_discord_notification(name, 'ユーザー追加', success=True, details={'idm': idm})
                    else:
                        flash(f'エラー: IDm "{idm}" は既にユーザー "{existing_user.name}" に登録されています。', 'danger')
            else:
                flash('エラー: IDm と名前は必須です。', 'danger')
        elif action == 'delete' and user_id:
            if password == DELETE_PASSWORD: # パスワードチェック
                with app.app_context():
                    user_to_delete = User.query.get(user_id)
                    if user_to_delete:
                        deleted_name = user_to_delete.name
                        AccessLog.query.filter_by(user_id=user_id).delete()
                        db.session.delete(user_to_delete)
                        db.session.commit()
                        flash(f'ユーザー "{deleted_name}" を削除しました。', 'success')
                        send_discord_notification(deleted_name, 'ユーザー削除', success=True)
                    else:
                        flash('エラー: ユーザーが見つかりませんでした。', 'danger')
            else:
                flash('エラー: パスワードが間違っています。', 'danger')
        # POSTリクエストのどのパスでも最終的にリダイレクトするように変更
        return redirect(url_for('manage_users'))

    # GETリクエストの場合
    users = User.query.all()
    return render_template('users.html', users=users)

@app.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):
    with app.app_context():
        user = User.query.get_or_404(user_id)

        if request.method == 'POST':
            old_name = user.name
            old_idm = user.idm
            new_name = request.form['name'].strip()
            password = request.form.get('password') # パスワードを取得

            if password == DELETE_PASSWORD: # パスワードチェック
                if new_name:
                    user.name = new_name
                    db.session.commit()
                    flash(f'ユーザー "{old_name}" の情報を更新しました。', 'success')
                    send_discord_notification(new_name, 'ユーザー更新', success=True, 
                                            details={'old_name': old_name, 'new_name': new_name, 'old_idm': old_idm, 'new_idm': user.idm})
                    return redirect(url_for('manage_users'))
                else:
                    flash('エラー: 名前は必須です。', 'danger')
            else:
                flash('エラー: パスワードが間違っています。', 'danger')
        
        return render_template('edit_user.html', user=user)

@app.route('/users/clear_logs/<int:user_id>', methods=['POST'])
def clear_user_logs(user_id):
    password = request.form.get('password') # パスワードを取得
    if password == DELETE_PASSWORD: # パスワードチェック
        with app.app_context():
            user = User.query.get(user_id)
            if user:
                num_deleted = AccessLog.query.filter_by(user_id=user.id).delete()
                db.session.commit()
                flash(f'ユーザー "{user.name}" の入退室ログ {num_deleted} 件を削除しました。', 'success')
                send_discord_notification(user.name, 'ログ削除', success=True, details={'deleted_count': num_deleted})
            else:
                flash('エラー: ログを削除するユーザーが見つかりませんでした。', 'danger')
    else:
        flash('エラー: パスワードが間違っています。', 'danger')
    return redirect(url_for('manage_users'))

@app.route('/ranking')
def show_ranking():
    current_year = datetime.now().year
    current_month = datetime.now().month

    year = request.args.get('year', type=int, default=current_year)
    month = request.args.get('month', type=int, default=current_month)

    if not (1 <= month <= 12):
        month = current_month
    
    with app.app_context():
        ranking = calculate_monthly_stay_time(year, month)
    
    return render_template('ranking.html', ranking=ranking, selected_year=year, selected_month=month, datetime=datetime)


# --- Excelファイルにログを記録する関数 ---
def update_excel_log():
    excel_file_path = "access_logs.xlsx"

    # データベースから全期間のログを取得
    with app.app_context():
        all_logs = AccessLog.query.order_by(AccessLog.timestamp).all()
        all_users = User.query.all()
    
    # ユーザー名とIDの紐づけ
    user_names = {user.id: user.name for user in all_users}

    # ユーザーごと、日ごとのセッションデータを集計
    user_daily_sessions = defaultdict(lambda: defaultdict(list))
    for log in all_logs:
        log_date = log.timestamp.date()
        user_daily_sessions[log.user_id][log_date].append(log)

    # Excelブックを新規作成
    workbook = Workbook()
    
    # ユーザーごとにシートを作成
    for user_id, daily_data in user_daily_sessions.items():
        user_name = user_names.get(user_id, f"不明なユーザー({user_id})")
        
        # シート名に使用できない文字を置き換える
        safe_sheet_name = user_name.replace('[', '').replace(']', '').replace(':', '').replace('/', '').replace('\\', '').replace('?', '').replace('*', '')
        
        # 既に存在しているデフォルトのシートを削除
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        # 新しいシートを作成
        sheet = workbook.create_sheet(title=safe_sheet_name)
        
        # ヘッダー行を書き込む
        sheet.append(["日付", "入室/退室記録", "合計滞在時間"])

        # 全期間のログから最初と最後の月を取得
        if all_logs:
            start_month = all_logs[0].timestamp.replace(day=1)
            end_month = all_logs[-1].timestamp.replace(day=1)
        else:
            # ログがない場合は今月を使用
            today = datetime.now()
            start_month = today.replace(day=1)
            end_month = today.replace(day=1)
        
        current_month = start_month
        while current_month <= end_month:
            # 月の最初の日と最後の日を取得
            first_day = current_month.date()
            _, last_day_of_month = calendar.monthrange(current_month.year, current_month.month)
            last_day = current_month.replace(day=last_day_of_month).date()

            # 月ごとのデータを書き込む
            sheet.append([f"--- {current_month.strftime('%Y年%m月')} ---", "", ""])
            
            # 1日から月末までループ
            for single_day_num in range(1, last_day_of_month + 1):
                current_day = current_month.replace(day=single_day_num).date()

                if current_day in daily_data:
                    # アクセスがあった日の処理
                    logs_on_day = daily_data[current_day]
                    in_out_times = []
                    total_stay_duration = timedelta(0)
                    entry_time = None
                    
                    for log in logs_on_day:
                        if log.status == '入室':
                            entry_time = log.timestamp
                        elif log.status == '退室' and entry_time:
                            duration = log.timestamp - entry_time
                            total_stay_duration += duration
                            in_out_times.append(f"{entry_time.strftime('%H:%M')}-{log.timestamp.strftime('%H:%M')}")
                            entry_time = None
                    
                    if entry_time:
                        in_out_times.append(f"{entry_time.strftime('%H:%M')}-未退室")
                        total_stay_duration += datetime.now() - entry_time
                    
                    # Excelの行データを作成
                    in_out_string = ", ".join(in_out_times)
                    total_seconds = total_stay_duration.total_seconds()
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    seconds = int(total_seconds % 60)
                    formatted_duration = f"{hours:02}:{minutes:02}:{seconds:02}"
                    
                    sheet.append([current_day.strftime('%Y-%m-%d'), in_out_string, formatted_duration])
                else:
                    # アクセスがなかった日の処理
                    sheet.append([current_day.strftime('%Y-%m-%d'), "", ""])

            # 次の月に進む
            if current_month.month == 12:
                current_month = current_month.replace(year=current_month.year + 1, month=1)
            else:
                current_month = current_month.replace(month=current_month.month + 1)

    # 最終的にExcelファイルを保存
    workbook.save(excel_file_path)
    print("Excelログを更新しました。")


# --- 定期実行タスク ---
last_auto_sign_out_date = None

def scheduled_system_notifications():
    """
    Excelログの更新と自動退室処理を定期的に実行するスレッド
    """
    global last_auto_sign_out_date
    while True:
        now = datetime.now()
        
        # --- 毎日23:59に自動退室処理を実行 ---
        if now.hour == 23 and now.minute == 59 and now.date() != last_auto_sign_out_date:
            with app.app_context():
                auto_sign_out()
            last_auto_sign_out_date = now.date()

        # --- Excelログの更新 ---
        # この処理も app_context 内で行う
        with app.app_context():
            update_excel_log()
            
        # 1分待機
        time.sleep(1 * 60)
        

def auto_sign_out():
    """
    退室忘れユーザーを自動的に退室させる関数
    毎日 23:59 に実行することを想定
    """
# --- auto_sign_out 関数を修正 ---
def auto_sign_out():
    """
    退室忘れユーザーを自動的に退室させる関数
    毎日 23:59 に実行することを想定
    """
    # この関数全体を `with app.app_context():` で囲む必要はない
    # なぜなら、呼び出し元で既に囲まれているから。
    subquery = db.session.query(
        AccessLog.user_id,
        db.func.max(AccessLog.timestamp).label('last_timestamp')
    ).group_by(AccessLog.user_id).subquery()

    users_to_sign_out = db.session.query(User).join(AccessLog).filter(
        AccessLog.user_id == subquery.c.user_id,
        AccessLog.timestamp == subquery.c.last_timestamp,
        AccessLog.status == '入室'
    ).all()


    if users_to_sign_out:
        now = datetime.now()
        # 誰かがいた場合のみ通知
        print(f"退室忘れのユーザーが{len(users_to_sign_out)}人いました。自動退室を記録します。")

        for user in users_to_sign_out:
            print(f"ユーザー '{user.name}' の自動退室を記録中...")

            # 新しい退室ログを作成
            log_entry = AccessLog(user_id=user.id, status='退室', timestamp=now)
            db.session.add(log_entry)
            db.session.commit()
            
            # Discordに個別の通知を送信
            send_discord_message(
                DISCORD_SYSTEM_MONITOR_WEBHOOK_URL,
                f"🚪 {user.name} さんの退室を、23:59に自動的に記録しました。",
                username="自動退室Bot"
            )
        # すべての処理が完了した後に、まとめて完了通知を送ることもできます
        # send_discord_message(
        #     DISCORD_SYSTEM_MONITOR_WEBHOOK_URL,
        #     f"✅ {len(users_to_sign_out)}人の自動退室処理が完了しました。",
        #     username="自動退室Bot"
        # )
    else:
        # 退室忘れのユーザーがいなかった場合は何も通知しない
        print("退室忘れのユーザーはいませんでした。")



# --- アプリケーション起動時の処理 ---
if __name__ == '__main__':
    with app.app_context():
        db.create_all()

        if not User.query.first():
            print("Adding initial users...")
            user1 = User(idm="F637CF05", name="Soma Taniguchi")
            db.session.add(user1)
            db.session.commit()
            print("Initial users added.")

    reader_thread = threading.Thread(target=card_reading_loop, daemon=True)
    reader_thread.start()
    
    excel_thread = threading.Thread(target=scheduled_system_notifications, daemon=True)
    excel_thread.start()

    try:
        app.run(host='0.0.0.0', port=5000, debug=False)
    except Exception as e:
        print(f"Flask App Error: {e}")
        sys.exit(1)
